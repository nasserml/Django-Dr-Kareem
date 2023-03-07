from django.shortcuts import render
from .forms import RegEmpForm

from openpyxl import load_workbook


def regEmp(request):
    
    m = ''
    
    form = RegEmpForm()
    
    if request.method == 'POST':
        
        form = RegEmpForm(request.POST)
        
        if form.is_valid():
            cd = form.cleaned_data
            wb = load_workbook(r'C:\Users\nasser\Documents\codes\python\django\emp\names.xlsx')
            
            ws = wb['emp']
            
            i = str(int(ws['F1'].value) + 1)
            
            ws['A' + i] = cd['name']
            ws['B' + i] = cd['salary']
            ws['C' + i] = cd['emp_email']
            ws['D' + i] = cd['address']
            
            ws['F1'] = i
            
            wb.save(r'C:\Users\nasser\Documents\codes\python\django\emp\names.xlsx')
            m = 'تم الحفظ بنجاح'
            
            form = RegEmpForm()
            
    
    return render(request, 'newemp.html', {'form' : form , 'output': m})